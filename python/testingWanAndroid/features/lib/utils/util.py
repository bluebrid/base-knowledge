import random
import string
from behave import *
from openpyxl import Workbook
from openpyxl import load_workbook
from config.config import settings
#http://zetcode.com/python/openpyxl/
#https://pynative.com/python-generate-random-string/
from itertools import zip_longest
class Utils(object):
    def __init__(self):
        object.__init__(self)

    def randomString(self, stringLength=10):
        """Generate a random string of fixed length """
        letters = string.ascii_lowercase
        return ''.join(random.choice(letters) for i in range(stringLength))

    def readExcel(self, sheetName):
        wb = load_workbook(settings["excel"])
        ws = wb.get_sheet_by_name(sheetName)
        lst = []
        header_row = None
        for i, row in enumerate(ws.iter_rows()):
            if header_row == None:
                header_row = [ c.internal_value for c in row ]
                continue
            row = dict(zip_longest(header_row, [ c.internal_value for c in row ]))
            lst.append(row)
        # for ele in lst:
        #     print('UserName:{}, Password:{}'.format(ele["username"],ele["password"]))
        return lst

    def writeExcel(self,sheetName, row):
        wb = load_workbook(settings["excel"])
        ws = wb.get_sheet_by_name(sheetName)
        ws.append(row)
        wb.save(settings["excel"])
        