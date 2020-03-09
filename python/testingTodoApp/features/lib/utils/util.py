import random
import string
from behave import *
from openpyxl import Workbook
from openpyxl import load_workbook
from config.config import settings
import logging
from logging import handlers
import os
#https://github.com/Nemo1122/AppiumDemo/blob/master/lib/utils.py
#http://zetcode.com/python/openpyxl/
#https://pynative.com/python-generate-random-string/
from itertools import zip_longest
class Utils(object):
    def __init__(self):
        object.__init__(self)

    def randomString(self, stringLength=10):
        """Generate a random string of fixed length """
        letters = string.ascii_lowercase
        return ''.join(random.choice(letters) for i in range(stringLength))

    def readExcel(self, sheetName):
        wb = load_workbook(settings["excel"])
        ws = wb.get_sheet_by_name(sheetName)
        lst = []
        header_row = None
        for i, row in enumerate(ws.iter_rows()):
            if header_row == None:
                header_row = [ c.internal_value for c in row ]
                continue
            row = dict(zip_longest(header_row, [ c.internal_value for c in row ]))
            lst.append(row)
        # for ele in lst:
        #     print('UserName:{}, Password:{}'.format(ele["username"],ele["password"]))
        return lst

    def writeExcel(self,sheetName, row):
        wb = load_workbook(settings["excel"])
        ws = wb.get_sheet_by_name(sheetName)
        ws.append(row)
        wb.save(settings["excel"])

# class MyLogger(object):
#     #https://github.com/Nemo1122/AppiumDemo/blob/master/lib/utils.py
#     def __init__(self,file_name,level='info',backCount=5,when='D'):
#         logger = logging.getLogger()
#         logger.setLevel(self.get_level(level))  # 设置日志的级别
#         # fl = logging.FileHandler(filename='a.log', mode='a', encoding='utf-8')
#         cl = logging.StreamHandler()  # 负责往控制台输出的
#         bl = handlers.TimedRotatingFileHandler(filename=file_name, when=when, interval=1,
#                                                backupCount=backCount, encoding='utf-8')
#         fmt = logging.Formatter('%(asctime)s - %(pathname)s[line:%(lineno)d]-%(levelname)s:%(message)s')
#         # 指定日志的格式
#         cl.setFormatter(fmt)  # 设置控制台输出的日志格式
#         bl.setFormatter(fmt)  # 设置文件里面写入的日志格式
#         logger.addHandler(cl)  # 把已经设置好的人放到办公室中
#         logger.addHandler(bl)  # 同上
#         self.logger = logger

#     def get_level(self, sss):
#         level = {
#             'debug': logging.DEBUG,
#             'info': logging.INFO,
#             'warn': logging.WARN,
#             'error': logging.ERROR
#         }
#         sss = sss.lower()
#         return level.get(sss)       

# # 目录
# PATH = os.path.dirname(
#     os.path.abspath(__file__)
# )

# REPORT_PATH = os.path.join(PATH, 'report')

# path = os.path.join(settings.LOG_PATH,  settings.LOG_NAME)  # 拼好日志的绝对路径
# log = MyLogger(path, settings.LOG_LEVEL).logger